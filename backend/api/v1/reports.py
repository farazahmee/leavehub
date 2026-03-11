"""
Reports / export routes
"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_
from datetime import date
from typing import Optional
import io
import csv


def _format_date_for_csv(val):
    """Format date for CSV so Excel displays it correctly (avoids ########)."""
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%d-%b-%Y")  # e.g. 15-Jan-2026
    return str(val)


from core.database import get_db
from core.permissions import require_admin_or_team_lead
from models.user import User
from models.employee import Employee
from models.attendance import Attendance
from models.leave import Leave
from models.payroll import Payroll

router = APIRouter()


@router.get("/attendance")
async def export_attendance(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    employee_id: Optional[int] = Query(None, description="Filter by employee ID"),
    employee_name: Optional[str] = Query(None, description="Filter by employee name (partial match)"),
    current_user: User = Depends(require_admin_or_team_lead),
    db: AsyncSession = Depends(get_db)
):
    """Export attendance as CSV. Tenant users see only their company's data."""
    query = select(Attendance, Employee).join(Employee, Attendance.employee_id == Employee.id)
    tenant_id = getattr(current_user, "tenant_id", None)
    if tenant_id is not None:
        query = query.where(Employee.tenant_id == tenant_id)
    if start_date:
        query = query.where(Attendance.date >= start_date)
    if end_date:
        query = query.where(Attendance.date <= end_date)
    if employee_id:
        query = query.where(Attendance.employee_id == employee_id)
    if employee_name and employee_name.strip():
        q = f"%{employee_name.strip()}%"
        query = query.where(
            or_(
                Employee.first_name.ilike(q),
                Employee.last_name.ilike(q),
                Employee.employee_id.ilike(q),
            )
        )
    query = query.order_by(Attendance.date.desc(), Attendance.check_in_time.desc())
    result = await db.execute(query)
    rows = result.all()

    output = io.StringIO()
    output.write("\ufeff")  # UTF-8 BOM for Excel compatibility
    writer = csv.writer(output)
    writer.writerow(["Employee ID", "Employee Name", "Date", "Check In", "Check Out", "Status"])
    for att, emp in rows:
        status = "Completed" if att.check_out_time else "Active"
        check_in = att.check_in_time.strftime("%H:%M") if att.check_in_time else ""
        check_out = att.check_out_time.strftime("%H:%M") if att.check_out_time else ""
        writer.writerow([
            emp.employee_id,
            f"{emp.first_name} {emp.last_name}",
            _format_date_for_csv(att.date),
            check_in,
            check_out,
            status,
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=attendance_export.csv"}
    )


@router.get("/leave")
async def export_leave(
    status_filter: Optional[str] = None,
    start_date: Optional[date] = Query(None, description="Filter leaves starting from this date"),
    end_date: Optional[date] = Query(None, description="Filter leaves ending before this date"),
    employee_id: Optional[int] = Query(None, description="Filter by employee ID"),
    employee_name: Optional[str] = Query(None, description="Filter by employee name (partial match)"),
    current_user: User = Depends(require_admin_or_team_lead),
    db: AsyncSession = Depends(get_db)
):
    """Export leave requests as XLSX. Tenant users see only their company's data."""
    from openpyxl import Workbook

    query = select(Leave, Employee).join(Employee, Leave.employee_id == Employee.id)
    tenant_id = getattr(current_user, "tenant_id", None)
    if tenant_id is not None:
        query = query.where(Employee.tenant_id == tenant_id)
    if status_filter:
        query = query.where(Leave.status == status_filter)
    if start_date:
        query = query.where(Leave.start_date >= start_date)
    if end_date:
        query = query.where(Leave.end_date <= end_date)
    if employee_id:
        query = query.where(Leave.employee_id == employee_id)
    if employee_name and employee_name.strip():
        q = f"%{employee_name.strip()}%"
        query = query.where(
            or_(
                Employee.first_name.ilike(q),
                Employee.last_name.ilike(q),
                Employee.employee_id.ilike(q),
            )
        )
    query = query.order_by(Leave.created_at.desc())
    result = await db.execute(query)
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave"
    headers = ["Employee ID", "Employee Name", "Leave Type", "Start Date", "End Date", "Days", "Status", "Reason"]
    ws.append(headers)
    date_format = "DD-MMM-YYYY"
    for leave, emp in rows:
        ws.append([
            emp.employee_id,
            f"{emp.first_name} {emp.last_name}",
            leave.leave_type or "",
            leave.start_date,
            leave.end_date,
            leave.days or "",
            leave.status or "",
            leave.reason or "",
        ])
    for row in range(2, len(rows) + 2):
        ws.cell(row=row, column=4).number_format = date_format
        ws.cell(row=row, column=5).number_format = date_format
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 6
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leave_export.xlsx"}
    )


@router.get("/payroll")
async def export_payroll(
    year: Optional[int] = None,
    month: Optional[int] = None,
    employee_id: Optional[int] = Query(None, description="Filter by employee ID"),
    employee_name: Optional[str] = Query(None, description="Filter by employee name (partial match)"),
    current_user: User = Depends(require_admin_or_team_lead),
    db: AsyncSession = Depends(get_db)
):
    """Export payroll as CSV. Tenant users see only their company's data."""
    query = select(Payroll, Employee).join(Employee, Payroll.employee_id == Employee.id)
    tenant_id = getattr(current_user, "tenant_id", None)
    if tenant_id is not None:
        query = query.where(Employee.tenant_id == tenant_id)
    if year:
        query = query.where(Payroll.year == year)
    if month:
        query = query.where(Payroll.month == month)
    if employee_id:
        query = query.where(Payroll.employee_id == employee_id)
    if employee_name and employee_name.strip():
        q = f"%{employee_name.strip()}%"
        query = query.where(
            or_(
                Employee.first_name.ilike(q),
                Employee.last_name.ilike(q),
                Employee.employee_id.ilike(q),
            )
        )
    query = query.order_by(Payroll.year.desc(), Payroll.month.desc())
    result = await db.execute(query)
    rows = result.all()

    output = io.StringIO()
    output.write("\ufeff")  # UTF-8 BOM for Excel compatibility
    writer = csv.writer(output)
    writer.writerow(["Employee ID", "Employee Name", "Month", "Year", "Basic", "Allowances", "Deductions", "Net"])
    for pay, emp in rows:
        writer.writerow([
            emp.employee_id,
            f"{emp.first_name} {emp.last_name}",
            pay.month,
            pay.year,
            pay.basic_salary or 0,
            pay.allowances or 0,
            pay.deductions or 0,
            pay.net_salary or 0,
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=payroll_export.csv"}
    )
